import os
import json
import uuid
import sqlite3
import random
from datetime import datetime
from io import BytesIO
from flask import Flask, request, jsonify, render_template_string, send_file, abort

try:
    from docx import Document
    from docx.oxml import parse_xml
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

app = Flask(__name__)

DB_PATH = "exam_system.db"

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def parse_docx_questions(file_stream):
    """Parse questions from a Word document."""
    if not DOCX_AVAILABLE:
        raise Exception("python-docx library not installed. Run: pip install python-docx")
    
    try:
        doc = Document(file_stream)
    except Exception as e:
        raise Exception(f"Failed to read Word file: {str(e)}")
    
    questions = []
    current_question = None
    current_options = []
    correct_answer = None
    marks = 1
    
    def extract_number(text):
        """Extract leading number from text like '1.', 'Question 1:', etc."""
        import re
        match = re.search(r'^(?:Question\s*)?(\d+)', text.strip())
        return match.group(1) if match else None
    
    def remove_question_number(text):
        """Remove leading question number from text like '1.', 'Question 1:', etc."""
        import re
        # Remove patterns like "1. ", "Question 1: ", "Q. ", etc.
        cleaned = re.sub(r'^(?:Question\s*)?(\d+)[.:\s]+', '', text.strip())
        return cleaned.strip()
    
    text_content = []
    for para in doc.paragraphs:
        text = para.text.strip()
        if text:
            text_content.append(text)
    
    i = 0
    while i < len(text_content):
        line = text_content[i]
        
        # Check if this is a question start (e.g., "1.", "Question 1:", etc.)
        if any(line.startswith(prefix) for prefix in ['1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.', '0.', 'Question ', 'Q.']):
            # Save previous question if exists
            if current_question:
                q_obj = {
                    'id': f'q{len(questions) + 1}',
                    'type': 'mcq' if current_options else 'short_answer',
                    'text': current_question,
                    'marks': marks
                }
                if current_options:
                    q_obj['options'] = current_options
                    q_obj['correctAnswer'] = str(correct_answer) if correct_answer is not None else '0'
                else:
                    q_obj['correctAnswer'] = correct_answer or ''
                questions.append(q_obj)
            
            # Start new question - remove question number
            current_question = remove_question_number(line)
            current_options = []
            correct_answer = None
            marks = 1
        
        # Check if this is an option line (A., B., C., D.)
        elif any(line.startswith(opt) for opt in ['A.', 'B.', 'C.', 'D.', 'a.', 'b.', 'c.', 'd.']):
            option_text = line[2:].strip() if len(line) > 2 else ''
            current_options.append(option_text)
        
        # Check for answer line
        elif line.lower().startswith('answer:') or line.lower().startswith('correct answer:'):
            answer_part = line.split(':', 1)[1].strip().upper()
            # Map A->0, B->1, C->2, D->3
            answer_map = {'A': 0, 'B': 1, 'C': 2, 'D': 3}
            correct_answer = answer_map.get(answer_part[0] if answer_part else 'A', 0)
        
        # Check for marks
        elif 'mark' in line.lower() and any(char.isdigit() for char in line):
            import re
            marks_match = re.search(r'(\d+)\s*marks?', line, re.IGNORECASE)
            if marks_match:
                marks = int(marks_match.group(1))
        
        i += 1
    
    # Don't forget the last question
    if current_question:
        q_obj = {
            'id': f'q{len(questions) + 1}',
            'type': 'mcq' if current_options else 'short_answer',
            'text': current_question,
            'marks': marks
        }
        if current_options:
            q_obj['options'] = current_options
            q_obj['correctAnswer'] = str(correct_answer) if correct_answer is not None else '0'
        else:
            q_obj['correctAnswer'] = correct_answer or ''
        questions.append(q_obj)
    
    return questions

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS exams (
            id TEXT PRIMARY KEY,
            teacher_email TEXT NOT NULL,
            title TEXT NOT NULL,
            duration INTEGER NOT NULL,
            questions TEXT NOT NULL,
            admin_token TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS sessions (
            id TEXT PRIMARY KEY,
            exam_id TEXT NOT NULL,
            student_name TEXT NOT NULL,
            student_id TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'in_progress',
            submitted_at TEXT,
            answers TEXT,
            score INTEGER DEFAULT 0,
            max_score INTEGER DEFAULT 0,
            question_order TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY (exam_id) REFERENCES exams(id)
        )
    """)
    cols = [r[1] for r in conn.execute("PRAGMA table_info(sessions)").fetchall()]
    if "score" not in cols:
        c.execute("ALTER TABLE sessions ADD COLUMN score INTEGER DEFAULT 0")
    if "max_score" not in cols:
        c.execute("ALTER TABLE sessions ADD COLUMN max_score INTEGER DEFAULT 0")
    if "question_order" not in cols:
        c.execute("ALTER TABLE sessions ADD COLUMN question_order TEXT")
    c.execute("""
        CREATE TABLE IF NOT EXISTS events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            exam_id TEXT NOT NULL,
            session_id TEXT NOT NULL,
            student_name TEXT NOT NULL,
            student_id TEXT NOT NULL,
            event_type TEXT NOT NULL,
            details TEXT,
            timestamp TEXT NOT NULL,
            FOREIGN KEY (exam_id) REFERENCES exams(id),
            FOREIGN KEY (session_id) REFERENCES sessions(id)
        )
    """)
    conn.commit()
    conn.close()

HOME_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>ExamGuard – Create Exam</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'Segoe UI', sans-serif; background: #f0f4f8; min-height: 100vh; }
  .header { background: #1a56db; color: #fff; padding: 18px 32px; display: flex; align-items: center; gap: 12px; }
  .header h1 { font-size: 1.4rem; font-weight: 700; }
  .header .sub { font-size: 0.85rem; opacity: 0.8; }
  .container { max-width: 820px; margin: 40px auto; padding: 0 20px; }
  .card { background: #fff; border-radius: 12px; box-shadow: 0 2px 12px rgba(0,0,0,0.08); padding: 32px; margin-bottom: 24px; }
  .card h2 { font-size: 1.1rem; color: #1e293b; margin-bottom: 20px; font-weight: 700; border-bottom: 2px solid #e2e8f0; padding-bottom: 12px; }
  label { display: block; font-size: 0.84rem; font-weight: 600; color: #374151; margin-bottom: 5px; margin-top: 14px; }
  input[type=text], input[type=email], input[type=number], textarea, select {
    width: 100%; padding: 10px 14px; border: 1.5px solid #d1d5db; border-radius: 8px;
    font-size: 0.94rem; outline: none; transition: border-color 0.2s;
  }
  input:focus, textarea:focus, select:focus { border-color: #1a56db; box-shadow: 0 0 0 3px rgba(26,86,219,0.1); }
  textarea { resize: vertical; min-height: 65px; }
  .btn { display: inline-flex; align-items: center; gap: 7px; padding: 10px 20px;
    border: none; border-radius: 8px; font-size: 0.88rem; font-weight: 700;
    cursor: pointer; transition: opacity 0.2s; }
  .btn-primary { background: #1a56db; color: #fff; }
  .btn-secondary { background: #e2e8f0; color: #374151; }
  .btn-danger { background: #ef4444; color: #fff; }
  .btn-success { background: #16a34a; color: #fff; }
  .btn:hover { opacity: 0.86; }
  .row2 { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }
  .upload-box { background: #f8fafc; border: 1.5px dashed #c7d2e0; border-radius: 10px; padding: 18px; margin-bottom: 18px; }
  .upload-box input[type=file] { margin-top: 8px; margin-bottom: 0; }
  .upload-help { font-style: italic; color: #6b7280; }
  .muted { font-size: 0.82rem; color: #94a3b8; font-weight: 500; }
  .question-block { background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 10px; padding: 18px; margin-top: 16px; }
  .q-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 14px; }
  .q-num { font-weight: 700; color: #1a56db; font-size: 0.88rem; }
  .option-row { display: flex; align-items: center; gap: 8px; margin-top: 8px; }
  .option-row input[type=text] { flex: 1; }
  #add-btns { display: flex; gap: 12px; margin-top: 18px; }
  .result-box { background: #f0fdf4; border: 1.5px solid #86efac; border-radius: 12px; padding: 24px; margin-top: 28px; }
  .result-box h3 { color: #15803d; margin-bottom: 16px; font-size: 1.05rem; font-weight: 700; }
  .link-row { display: flex; align-items: center; gap: 10px; background: #fff;
    border: 1px solid #d1fae5; border-radius: 8px; padding: 11px 16px; margin-top: 10px; }
  .link-row .link-label { font-size: 0.75rem; color: #16a34a; min-width: 90px; font-weight: 700; white-space: nowrap; }
  .link-row input { flex: 1; border: none; outline: none; font-size: 0.88rem; color: #1e293b; background: transparent; }
  .copy-btn { background: #1a56db; color: #fff; border: none; border-radius: 6px; padding: 6px 12px; font-size: 0.78rem; cursor: pointer; white-space: nowrap; }
  .toast { position: fixed; bottom: 24px; right: 24px; background: #1e293b; color: #fff;
    padding: 12px 22px; border-radius: 8px; font-size: 0.88rem; display: none; z-index: 9999; box-shadow: 0 4px 12px rgba(0,0,0,0.15); }
  .create-btn { width: 100%; padding: 15px; font-size: 1rem; border-radius: 10px; margin-top: 4px; }
</style>
</head>
<body>
<div class="header">
  <svg width="30" height="30" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2">
    <path stroke-linecap="round" stroke-linejoin="round" d="M9 12l2 2 4-4m5.618-4.016A11.955 11.955 0 0112 2.944a11.955 11.955 0 01-8.618 3.04A12.02 12.02 0 003 9c0 5.591 3.824 10.29 9 11.622 5.176-1.332 9-6.03 9-11.622 0-1.042-.133-2.052-.382-3.016z"/>
  </svg>
  <div>
    <h1>ExamGuard</h1>
    <div class="sub">Online Exam Anti-Cheating System</div>
  </div>
</div>

<div class="container">
  <div class="card">
    <h2>📋 Exam Details</h2>
    <div class="row2">
      <div>
        <label>Your Gmail Address OR Name</label>
        <input type="email" id="teacher-email" placeholder="teacher@gmail.com or teacher" />
      </div>
      <div>
        <label>Exam Title</label>
        <input type="text" id="exam-title" placeholder="e.g. Midterm Mathematics" />
      </div>
    </div>
    <div style="max-width:200px">
      <label>Duration (minutes)</label>
      <input type="number" id="exam-duration" value="60" min="1" max="300" />
    </div>
  </div>

  <div class="card">
    <h2>❓ Questions</h2>
    <div id="questions-container"></div>
    <div id="add-btns" style="display:flex;gap:12px;margin-top:18px;flex-wrap:wrap;align-items:center">
      <button class="btn btn-primary" onclick="addQuestion('mcq')">+ Add MCQ</button>
      <button class="btn btn-secondary" onclick="addQuestion('short')">+ Add Short Answer</button>
      <span style="margin-left:12px;border-left:1px solid #d1d5db;padding-left:12px;display:flex;gap:8px;align-items:center;flex-wrap:wrap">
        <input type="file" id="word-file" accept=".docx" style="padding:6px;font-size:0.76rem;width:auto" />
        <button type="button" class="btn btn-secondary" style="padding:8px 14px;font-size:0.82rem" onclick="importWordFile()">📄 Import Word</button>
        <span class="muted" id="import-status" style="font-size:0.76rem">No file</span>
      </span>
    </div>
    <p class="upload-help" style="margin-top:12px;font-size:0.76rem;color:#64748b;line-height:1.5">
      <strong>Word format:</strong> Questions: "1." or "Question 1" | Options: "A." "B." "C." "D." | Answer: "Answer: A" | Marks: "Marks: 2" (optional)
    </p>
  </div>

  <button class="btn btn-success create-btn" onclick="createExam()">🚀 Create Exam &amp; Generate Links</button>

  <div id="result-section" style="display:none">
    <div class="result-box">
      <h3>✅ Exam Created Successfully!</h3>
      <p style="color:#374151;font-size:0.88rem;margin-bottom:12px">Share the student link with your students and keep the admin link private.</p>
      <div class="link-row">
        <span class="link-label">📎 Student Link</span>
        <input type="text" id="exam-link" readonly />
        <button class="copy-btn" onclick="copyLink('exam-link')">Copy</button>
      </div>
      <div class="link-row" style="margin-top:8px">
        <span class="link-label">🔒 Admin Link</span>
        <input type="text" id="admin-link" readonly />
        <button class="copy-btn" onclick="copyLink('admin-link')">Copy</button>
      </div>
    </div>
  </div>
</div>
<div class="toast" id="toast">Copied to clipboard!</div>

<script>
let qCount = 0;

function addQuestion(type, preset = null) {
  qCount++;
  const n = qCount;
  const container = document.getElementById('questions-container');
  const div = document.createElement('div');
  div.className = 'question-block';
  div.id = 'q-' + n;
  div.dataset.type = type;

  let optHtml = '';
  if (type === 'mcq') {
    const options = (preset && Array.isArray(preset.options) && preset.options.length ? preset.options : ['', '', '', '']).slice(0, 4);
    while (options.length < 4) options.push('');
    const selected = preset && preset.correctAnswer !== undefined ? String(preset.correctAnswer) : '';
    optHtml = `<div style="margin-top:10px">
      <label style="margin-top:0">Options <small style="font-weight:400;color:#888">(select the correct answer)</small></label>
      ${options.map((value, idx) => `
        <div class="option-row">
          <input type="radio" name="correct-${n}" value="${idx + 1}" id="r-${n}-${idx + 1}" ${selected === String(idx) ? 'checked' : ''}>
          <input type="text" id="opt-${n}-${idx + 1}" placeholder="Option ${idx + 1}" value="${escapeHtml(value)}" />
        </div>`).join('')}
    </div>`;
  } else if (type === 'short') {
    optHtml = `<div style="margin-top:10px">
      <label style="margin-top:0">Correct Answer <small style="font-weight:400;color:#888">(for auto-grading)</small></label>
      <input type="text" id="correct-${n}" placeholder="Expected answer" value="${escapeHtml((preset && preset.correctAnswer) || '')}" />
    </div>`;
  }

  const marksValue = preset && preset.marks !== undefined ? preset.marks : 1;
  const marksHtml = `<div style="margin-top:12px;display:flex;gap:12px;align-items:flex-end;">
      <div style="flex:1">
        <label>Marks</label>
        <input type="number" id="marks-${n}" value="${escapeHtml(String(marksValue))}" min="0" step="1" />
      </div>
    </div>`;

  div.innerHTML = `
    <div class="q-header">
      <span class="q-num">Q${n} — ${type === 'mcq' ? 'Multiple Choice' : 'Short Answer'}</span>
      <button class="btn btn-danger" style="padding:5px 11px;font-size:0.76rem" onclick="document.getElementById('q-${n}').remove()">Remove</button>
    </div>
    <label style="margin-top:0">Question Text</label>
    <textarea id="qtext-${n}" placeholder="Enter your question here…">${escapeHtml((preset && preset.text) || '')}</textarea>
    ${marksHtml}
    ${optHtml}`;
  container.appendChild(div);
}

function escapeHtml(value) {
  return String(value || '')
    .replace(/&/g, '&amp;')
    .replace(/</g, '&lt;')
    .replace(/>/g, '&gt;')
    .replace(/"/g, '&quot;');
}

function loadImportedQuestions(questions) {
  for (const q of questions) {
    addQuestion(q.type === 'short_answer' ? 'short' : 'mcq', q);
  }
}

async function importWordFile() {
  const fileInput = document.getElementById('word-file');
  const status = document.getElementById('import-status');
  if (!fileInput.files.length) {
    alert('Please choose a Word (.docx) file first.');
    return;
  }
  
  const file = fileInput.files[0];
  const formData = new FormData();
  formData.append('file', file);
  
  status.textContent = 'Parsing...';
  status.style.color = '#0ea5e9';
  
  try {
    const res = await fetch('/exam-api/parse-docx', {
      method: 'POST',
      body: formData
    });
    
    const data = await res.json();
    if (!res.ok) {
      alert('Error: ' + (data.error || 'Failed to parse Word file'));
      status.textContent = 'Failed to import.';
      status.style.color = '#ef4444';
      return;
    }
    
    if (!data.questions || data.questions.length === 0) {
      alert('No questions found in the Word file.');
      status.textContent = 'No questions extracted.';
      status.style.color = '#f59e0b';
      return;
    }
    
    // Clear existing questions
    document.getElementById('questions-container').innerHTML = '';
    qCount = 0;
    
    // Load imported questions
    loadImportedQuestions(data.questions);
    
    status.textContent = `✓ ${data.questions.length} questions imported successfully!`;
    status.style.color = '#10b981';
  } catch (err) {
    alert('Error importing file: ' + err.message);
    status.textContent = 'Error importing file.';
    status.style.color = '#ef4444';
  }
}

function gatherQuestions() {
  const blocks = document.querySelectorAll('.question-block');
  const qs = [];
  for (const b of blocks) {
    const n = b.id.replace('q-', '');
    const type = b.dataset.type;
    const text = document.getElementById('qtext-' + n).value.trim();
    if (!text) continue;
    const q = { id: 'q' + n, type: type === 'mcq' ? 'mcq' : 'short_answer', text };
    const marksEl = document.getElementById('marks-' + n);
    q.marks = Math.max(0, parseInt(marksEl?.value) || 0);
    if (type === 'mcq') {
      q.options = [];
      let correct = '';
      for (let i = 1; i <= 4; i++) {
        const el = document.getElementById('opt-' + n + '-' + i);
        if (el) q.options.push(el.value.trim());
        const r = document.getElementById('r-' + n + '-' + i);
        if (r && r.checked) correct = String(i - 1);
      }
      q.options = q.options.filter(Boolean);
      q.correctAnswer = correct;
    } else {
      const correctEl = document.getElementById('correct-' + n);
      q.correctAnswer = (correctEl ? correctEl.value.trim() : '');
    }
    qs.push(q);
  }
  return qs;
}

async function createExam() {
  const email = document.getElementById('teacher-email').value.trim();
  const title = document.getElementById('exam-title').value.trim();
  const duration = parseInt(document.getElementById('exam-duration').value) || 60;
  if (!email || !title) { alert('Please enter your Gmail and exam title.'); return; }
  const questions = gatherQuestions();
  if (questions.length === 0) { alert('Please add at least one question.'); return; }

  const res = await fetch('/exam-api/exams', {
    method: 'POST',
    headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify({ teacherEmail: email, title, duration, questions })
  });
  const data = await res.json();
  if (!res.ok) { alert(data.error || 'Failed to create exam.'); return; }

  document.getElementById('exam-link').value = data.examLink;
  document.getElementById('admin-link').value = data.adminLink;
  document.getElementById('result-section').style.display = 'block';
  document.getElementById('result-section').scrollIntoView({ behavior: 'smooth' });
}

function copyLink(id) {
  const input = document.getElementById(id);
  const value = input.value || '';
  if (!value) {
    alert('Nothing to copy yet. Create an exam first.');
    return;
  }

  const showToast = () => {
    const t = document.getElementById('toast');
    t.style.display = 'block';
    setTimeout(() => t.style.display = 'none', 2000);
  };

  const fallbackCopy = () => {
    input.select();
    input.setSelectionRange(0, 99999); // For mobile
    try {
      document.execCommand('copy');
      showToast();
    } catch (err) {
      alert('Copy failed. Please select the link and press Ctrl+C (or Cmd+C on Mac).');
    }
  };

  if (navigator.clipboard && navigator.clipboard.writeText) {
    navigator.clipboard.writeText(value)
      .then(showToast)
      .catch(() => fallbackCopy());
  } else {
    fallbackCopy();
  }
}
</script>
</body>
</html>
"""

REGISTER_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{{ exam.title }} – Student Registration</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'Segoe UI', sans-serif; background: linear-gradient(135deg,#1a56db 0%,#1e40af 100%);
    min-height: 100vh; display: flex; align-items: center; justify-content: center; padding: 20px; }
  .card { background: #fff; border-radius: 16px; box-shadow: 0 20px 60px rgba(0,0,0,0.2); padding: 40px; max-width: 440px; width: 100%; }
  .logo { display: flex; align-items: center; gap: 10px; margin-bottom: 28px; }
  .logo h1 { font-size: 1.3rem; color: #1a56db; font-weight: 700; }
  h2 { font-size: 1.4rem; color: #1e293b; font-weight: 700; margin-bottom: 6px; }
  .meta { color: #64748b; font-size: 0.88rem; margin-bottom: 24px; }
  .meta .pill { background: #eff6ff; color: #1a56db; border-radius: 6px; padding: 3px 9px; font-weight: 600; }
  label { display: block; font-size: 0.84rem; font-weight: 600; color: #374151; margin-bottom: 6px; margin-top: 16px; }
  input { width: 100%; padding: 12px 14px; border: 1.5px solid #d1d5db; border-radius: 8px; font-size: 0.95rem; outline: none; transition: border-color 0.2s; }
  input:focus { border-color: #1a56db; box-shadow: 0 0 0 3px rgba(26,86,219,0.1); }
  .warning { background: #fef9c3; border: 1px solid #fde047; border-radius: 8px; padding: 12px 16px; margin-top: 20px; font-size: 0.82rem; color: #713f12; line-height: 1.5; }
  .warning ul { margin-left: 18px; margin-top: 6px; }
  .btn { display: block; width: 100%; padding: 14px; background: #1a56db; color: #fff;
    border: none; border-radius: 10px; font-size: 1rem; font-weight: 700; cursor: pointer; margin-top: 22px; transition: opacity 0.2s; }
  .btn:hover { opacity: 0.88; }
</style>
</head>
<body>
<div class="card">
  <div class="logo">
    <svg width="26" height="26" fill="none" viewBox="0 0 24 24" stroke="#1a56db" stroke-width="2">
      <path stroke-linecap="round" stroke-linejoin="round" d="M9 12l2 2 4-4m5.618-4.016A11.955 11.955 0 0112 2.944a11.955 11.955 0 01-8.618 3.04A12.02 12.02 0 003 9c0 5.591 3.824 10.29 9 11.622 5.176-1.332 9-6.03 9-11.622 0-1.042-.133-2.052-.382-3.016z"/>
    </svg>
    <h1>ExamGuard</h1>
  </div>
  <h2>{{ exam.title }}</h2>
  <p class="meta">Duration: <span class="pill">{{ exam.duration }} min</span></p>

  <div class="warning">
    ⚠️ <strong>This exam is monitored:</strong>
    <ul>
      <li>Tab switching and window changes are tracked</li>
      <li>Copy/paste and keyboard shortcuts are blocked</li>
      <li>More than one violation will auto-terminate your exam and block future access</li>
    </ul>
  </div>

  <label>Full Name</label>
  <input type="text" id="student-name" placeholder="Enter your full name" />
  <label>Student ID</label>
  <input type="text" id="student-id" placeholder="Enter your student ID" />

  <button class="btn" onclick="startExam()">Start Exam →</button>
</div>
<script>
async function startExam() {
  const name = document.getElementById('student-name').value.trim();
  const sid = document.getElementById('student-id').value.trim();
  if (!name || !sid) { alert('Please enter your name and student ID.'); return; }
  const res = await fetch('/exam-api/exams/{{ exam.id }}/students', {
    method: 'POST',
    headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify({ name, studentId: sid })
  });
  const data = await res.json();
  if (!res.ok) { alert(data.error || 'Registration failed.'); return; }
  window.location.href = '/exam/{{ exam.id }}/take?sessionId=' + data.sessionId;
}
</script>
</body>
</html>
"""

EXAM_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{{ exam.title }} – Exam in Progress</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'Segoe UI', sans-serif; background: #0f172a; color: #e2e8f0; user-select: none; min-height: 100vh; }
  .topbar { background: #1e3a8a; padding: 12px 20px; display: flex; align-items: center;
    justify-content: space-between; position: sticky; top: 0; z-index: 200; border-bottom: 2px solid #1a56db; }
  .topbar-left { display: flex; align-items: center; gap: 12px; }
  .topbar h1 { font-size: 0.95rem; font-weight: 700; color: #fff; }
  .timer { background: #1a56db; color: #fff; border-radius: 8px; padding: 7px 18px;
    font-size: 1.1rem; font-weight: 800; font-variant-numeric: tabular-nums; letter-spacing: 1px; }
  .timer.warn { background: #dc2626; animation: pulse 1s infinite; }
  @keyframes pulse { 0%,100%{opacity:1} 50%{opacity:0.7} }
  #vbar { display: none; background: #fef3c7; border-bottom: 2px solid #f59e0b;
    padding: 7px 20px; font-size: 0.82rem; color: #92400e; font-weight: 600; }
  .main { display: grid; grid-template-columns: 1fr; gap: 0; min-height: calc(100vh - 55px); }
  @media (min-width: 769px) { .main { grid-template-columns: 1fr 320px; } }
  .questions-col { padding: 20px; overflow-y: auto; }
  .student-banner { background: #1e3a8a; border: 1px solid #1d4ed8; border-radius: 10px;
    padding: 12px 18px; margin-bottom: 18px; display: flex; gap: 20px; font-size: 0.84rem; color: #93c5fd; }
  .student-banner strong { color: #fff; }
  .q-card { background: #1e293b; border: 1px solid #334155; border-radius: 10px;
    padding: 20px; margin-bottom: 14px; }
  .q-num { color: #60a5fa; font-size: 0.75rem; font-weight: 700; margin-bottom: 8px; }
  .q-text { font-size: 0.97rem; color: #f1f5f9; margin-bottom: 14px; line-height: 1.5; font-weight: 500; }
  .options label { display: flex; align-items: center; gap: 10px; padding: 9px 14px;
    border: 1px solid #334155; border-radius: 8px; margin-bottom: 7px; cursor: pointer;
    color: #cbd5e1; font-size: 0.9rem; transition: all 0.15s; }
  .options label:hover { background: #1d4ed8; border-color: #3b82f6; color: #fff; }
  .options label:has(input:checked) { background: #1d4ed8; border-color: #60a5fa; color: #fff; }
  textarea.answer { width: 100%; background: #0f172a; border: 1.5px solid #334155; border-radius: 8px;
    color: #f1f5f9; padding: 10px 14px; font-size: 0.9rem; resize: vertical; min-height: 90px;
    outline: none; font-family: inherit; }
  textarea.answer:focus { border-color: #3b82f6; }
  .submit-area { padding: 0 20px 28px; }
  .submit-btn { width: 100%; padding: 15px; background: #16a34a; color: #fff; border: none;
    border-radius: 10px; font-size: 1rem; font-weight: 700; cursor: pointer; transition: opacity 0.2s; }
  .submit-btn:hover { opacity: 0.88; }
  .monitor-sidebar { background: #0f172a; border-left: 1px solid #1e293b; padding: 16px; display: flex; flex-direction: column; gap: 14px; }
  @media (max-width: 768px) { .monitor-sidebar { border-left: none; border-top: 1px solid #1e293b; padding: 14px; } }
  .monitor-title { color: #94a3b8; font-size: 0.75rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.8px; }
  .monitor-stats { background: #1e293b; border-radius: 8px; padding: 12px; font-size: 0.78rem; }
  .monitor-stats .row { display: flex; justify-content: space-between; color: #94a3b8; padding: 3px 0; }
  .monitor-stats .row strong { color: #f1f5f9; }
  .monitor-note { background: #1e293b; border: 1px solid #334155; border-radius: 8px; padding: 10px; font-size: 0.76rem; color: #64748b; line-height: 1.5; }
  .violations-log { flex: 1; overflow-y: auto; }
  .violations-log .log-title { color: #94a3b8; font-size: 0.75rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.8px; margin-bottom: 8px; }
  .log-entry { background: #1e293b; border-left: 3px solid #ef4444; border-radius: 6px; padding: 8px 10px; margin-bottom: 6px; font-size: 0.76rem; }
  .log-entry .log-type { color: #fca5a5; font-weight: 700; }
  .log-entry .log-time { color: #64748b; float: right; }
  .overlay { display: none; position: fixed; inset: 0; background: rgba(0,0,0,0.85);
    z-index: 1000; align-items: center; justify-content: center; }
  .overlay.active { display: flex; }
  .overlay-box { background: #fff; border-radius: 16px; padding: 40px; text-align: center; max-width: 400px; width: 90%; }
  .overlay-box .icon { font-size: 3.5rem; margin-bottom: 16px; }
  .overlay-box h2 { color: #1e293b; font-size: 1.4rem; margin-bottom: 10px; }
  .overlay-box.terminated h2 { color: #dc2626; }
  .overlay-box p { color: #64748b; font-size: 0.9rem; line-height: 1.6; }
</style>
</head>
<body>
<div id="vbar"></div>
<div class="topbar">
  <div class="topbar-left">
    <svg width="22" height="22" fill="none" viewBox="0 0 24 24" stroke="#60a5fa" stroke-width="2">
      <path stroke-linecap="round" stroke-linejoin="round" d="M9 12l2 2 4-4m5.618-4.016A11.955 11.955 0 0112 2.944a11.955 11.955 0 01-8.618 3.04A12.02 12.02 0 003 9c0 5.591 3.824 10.29 9 11.622 5.176-1.332 9-6.03 9-11.622 0-1.042-.133-2.052-.382-3.016z"/>
    </svg>
    <h1>{{ exam.title }}</h1>
  </div>
  <div class="timer" id="timer">{{ exam.duration }}:00</div>
</div>
<div class="main">
  <div class="questions-col">
    <div class="student-banner">
      <div>Student: <strong id="sname">—</strong></div>
      <div>ID: <strong id="sid">—</strong></div>
      <div style="margin-left:auto;font-size:0.76rem;color:#64748b">🔴 Exam being monitored</div>
    </div>
    <form id="exam-form">
    {% for q in questions %}
    <div class="q-card">
      <div class="q-num">Q{{ loop.index }} / {{ questions|length }} — {% if q.type == 'mcq' %}Multiple Choice{% else %}Short Answer{% endif %}</div>
      <div class="q-text">{{ q.text }}</div>
      {% if q.type == 'mcq' %}
      <div class="options">
        {% for opt in q.options %}
        <label>
          <input type="radio" name="{{ q.id }}" value="{{ loop.index0 }}" />
          <span>{{ opt }}</span>
        </label>
        {% endfor %}
      </div>
      {% else %}
      <textarea class="answer" name="{{ q.id }}" placeholder="Write your answer here…"></textarea>
      {% endif %}
    </div>
    {% endfor %}
    </form>
  </div>
  <div class="monitor-sidebar">
    <div class="monitor-title">🚨 Exam Monitoring</div>
    <div class="monitor-stats">
      <div class="row"><span>Violations</span><strong id="stat-vcount" style="color:#ef4444">0</strong></div>
    </div>
    <div class="monitor-note">⚠️ This exam is monitored for suspicious activity like tab switching, copy/paste, and system shortcuts. More than one violation will automatically terminate the exam and block future access.</div>
    <div class="violations-log">
      <div class="log-title">Recent Violations</div>
      <div id="log-list"><div style="color:#475569;font-size:0.76rem">No violations logged yet.</div></div>
    </div>
  </div>
</div>
<div class="submit-area">
  <button class="submit-btn" onclick="submitExam(false)">✅ Submit Exam</button>
</div>
<div class="overlay" id="submit-overlay">
  <div class="overlay-box">
    <div class="icon">✅</div>
    <h2>Exam Submitted</h2>
    <p>Your answers have been recorded successfully. You may now close this window.</p>
  </div>
</div>
<script>
const EXAM_ID = "{{ exam.id }}";
const SESSION_ID = new URLSearchParams(window.location.search).get('sessionId') || '';
const IS_MOBILE = window.innerWidth < 768;
let violated = {}, totalViolations = 0, terminated = false, submitted = false;
let timerSec = {{ exam.duration }} * 60;
let timerInterval = null;

fetch('/exam-api/session/' + SESSION_ID).then(r=>r.json()).then(d=>{
  if(d.studentName) document.getElementById('sname').textContent=d.studentName;
  if(d.studentId) document.getElementById('sid').textContent=d.studentId;
}).catch(()=>{});

if (!IS_MOBILE) {
  function tryFullscreen() {
    const el=document.documentElement;
    (el.requestFullscreen||el.webkitRequestFullscreen||el.mozRequestFullScreen||function(){}).call(el);
  }
  document.addEventListener('click', tryFullscreen, {once:true});
  tryFullscreen();
  document.addEventListener('fullscreenchange', ()=>{
    if(!document.fullscreenElement&&!terminated&&!submitted)
      logEvent('FULLSCREEN_EXIT','Student exited fullscreen mode');
  });
}

function tickTimer() {
  if(terminated||submitted) return;
  timerSec--;
  if(timerSec<=0){submitExam(false,'Time expired');return;}
  const m=String(Math.floor(timerSec/60)).padStart(2,'0');
  const s=String(timerSec%60).padStart(2,'0');
  const el=document.getElementById('timer');
  el.textContent=m+':'+s;
  if(timerSec<300) el.classList.add('warn');
}
timerInterval=setInterval(tickTimer,1000);

document.addEventListener('keydown',e=>{
  if(terminated) return;
  const ctrl=e.ctrlKey||e.metaKey, k=e.key.toLowerCase();
  if(ctrl&&['c','v','u','s','p','a','x'].includes(k)){e.preventDefault();logEvent('KEY_BLOCKED','Ctrl+'+k.toUpperCase()+' blocked');return;}
  if(e.key==='PrintScreen'){e.preventDefault();logEvent('PRINTSCREEN','Print Screen blocked');return;}
  if(e.key==='Meta'||e.key==='OS'){e.preventDefault();logEvent('WINDOWS_KEY','Windows key blocked');return;}
  if(e.key==='F12'){e.preventDefault();logEvent('DEVTOOLS_ATTEMPT','F12 blocked');return;}
  if(ctrl&&e.shiftKey){e.preventDefault();logEvent('COPILOT_ATTEMPT','Ctrl+Shift blocked');return;}
  if(e.altKey&&e.key==='Tab'){e.preventDefault();logEvent('ALT_TAB','Alt+Tab blocked');return;}
  if(e.key==='Escape'&&!IS_MOBILE) e.preventDefault();
});
document.addEventListener('contextmenu',e=>{if(!terminated){e.preventDefault();logEvent('RIGHT_CLICK','Right-click blocked');}});
document.addEventListener('copy',e=>{if(!terminated){e.preventDefault();logEvent('COPY_ATTEMPT','Copy blocked');}});
document.addEventListener('paste',e=>{if(!terminated){e.preventDefault();logEvent('PASTE_ATTEMPT','Paste blocked');}});
document.addEventListener('cut',e=>{if(!terminated){e.preventDefault();logEvent('CUT_ATTEMPT','Cut blocked');}});
document.addEventListener('visibilitychange',()=>{if(document.hidden&&!terminated) logEvent('TAB_SWITCH','Tab switched or window minimized');});
window.addEventListener('blur',()=>{if(!terminated) logEvent('WINDOW_BLUR','Exam window lost focus');});
setInterval(()=>{if(terminated) return;if(window.outerWidth-window.innerWidth>160||window.outerHeight-window.innerHeight>160) logEvent('DEVTOOLS_OPEN','DevTools likely open');},5000);

async function logEvent(type, details='') {
  if(terminated) return;
  violated[type]=(violated[type]||0)+1;
  totalViolations++;
  document.getElementById('stat-vcount').textContent=totalViolations;
  const vbar=document.getElementById('vbar');
  vbar.textContent='⚠️ '+type+' — '+details;
  vbar.style.display='block';
  setTimeout(()=>vbar.style.display='none',4000);
  const logList=document.getElementById('log-list');
  const now=new Date().toLocaleTimeString();
  const entry=document.createElement('div');
  entry.className='log-entry';
  entry.innerHTML='<span class="log-type">'+type+'</span><span class="log-time">'+now+'</span><br><span style="color:#94a3b8">'+details+'</span>';
  if(logList.firstChild&&logList.firstChild.textContent==='No violations logged yet.') logList.innerHTML='';
  logList.insertBefore(entry,logList.firstChild);
  try {
    const res=await fetch('/exam-api/events',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({examId:EXAM_ID,sessionId:SESSION_ID,eventType:type,details})});
    const data=await res.json();
    if(data.terminate&&!terminated) terminateExam('Repeated violation: '+type);
  } catch(e){}
}

async function submitExam(isTerminated=false, reason='') {
  if(submitted||terminated) return;
  submitted=true;
  clearInterval(timerInterval);
  const form=document.getElementById('exam-form');
  const answers={};
  for(const el of form.elements){
    if(!el.name) continue;
    if(el.type==='radio'&&el.checked) answers[el.name]=el.value;
    if(el.tagName==='TEXTAREA') answers[el.name]=el.value;
  }
  await fetch('/exam-api/exams/'+EXAM_ID+'/submit',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({sessionId:SESSION_ID,answers,terminated:isTerminated})}).catch(()=>{});
  // Always show submitted overlay
  document.getElementById('submit-overlay').classList.add('active');
}

function terminateExam(reason='') {
  if(terminated) return;
  terminated=true;
  submitExam(true,reason);
}

</script>
</body>
</html>
"""


ADMIN_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Admin – {{ exam.title }}</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'Segoe UI', sans-serif; background: #f0f4f8; }
  .header { background: #1a56db; color: #fff; padding: 16px 28px; display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 12px; }
  .header-left h1 { font-size: 1.05rem; font-weight: 700; }
  .header-left .meta { font-size: 0.78rem; opacity: 0.8; margin-top: 3px; }
  .header-right { display: flex; gap: 10px; align-items: center; }
  .container { max-width: 1200px; margin: 24px auto; padding: 0 20px; }
  .stats { display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 14px; margin-bottom: 22px; }
  .stat { background: #fff; border-radius: 10px; padding: 18px; box-shadow: 0 1px 5px rgba(0,0,0,0.07); text-align: center; }
  .stat .num { font-size: 2rem; font-weight: 800; color: #1a56db; }
  .stat .lbl { font-size: 0.76rem; color: #64748b; margin-top: 3px; }
  .card { background: #fff; border-radius: 10px; box-shadow: 0 1px 5px rgba(0,0,0,0.07); padding: 22px; margin-bottom: 22px; }
  .card-hdr { display: flex; align-items: center; justify-content: space-between; margin-bottom: 14px; }
  .card-hdr h2 { font-size: 0.95rem; font-weight: 700; color: #1e293b; }
  .card-hdr .note { font-size: 0.76rem; color: #94a3b8; }
  table { width: 100%; border-collapse: collapse; font-size: 0.83rem; }
  th { text-align: left; padding: 9px 12px; background: #f8fafc; color: #374151; font-weight: 700; border-bottom: 2px solid #e2e8f0; }
  td { padding: 9px 12px; border-bottom: 1px solid #f1f5f9; color: #1e293b; }
  tr:hover td { background: #f8fafc; }
  .badge { display: inline-block; padding: 2px 9px; border-radius: 100px; font-size: 0.73rem; font-weight: 700; }
  .bg-green{background:#dcfce7;color:#15803d} .bg-red{background:#fee2e2;color:#dc2626}
  .bg-yellow{background:#fef9c3;color:#92400e} .bg-blue{background:#dbeafe;color:#1d4ed8} .bg-gray{background:#f1f5f9;color:#64748b}
  .btn { display: inline-flex; align-items: center; gap: 6px; padding: 8px 16px;
    border: none; border-radius: 7px; font-size: 0.83rem; font-weight: 700; cursor: pointer; transition: opacity 0.2s; }
  .btn-green{background:#16a34a;color:#fff} .btn-blue{background:#1a56db;color:#fff}
  .btn:hover { opacity: 0.86; }
  .mono { font-family: monospace; font-size: 0.78rem; }
</style>
</head>
<body>
<div class="header">
  <div class="header-left">
    <h1>🛡️ Admin Dashboard — {{ exam.title }}</h1>
    <div class="meta">Teacher: {{ exam.teacher_email }} &nbsp;·&nbsp; Duration: {{ exam.duration }} min &nbsp;·&nbsp; ID: {{ exam.id }}</div>
  </div>
  <div class="header-right">
    <button class="btn btn-green" onclick="downloadXlsx()">⬇️ Download Excel (.xlsx)</button>
    <button class="btn btn-blue" onclick="loadDashboard()">🔄 Refresh</button>
  </div>
</div>
<div class="container">
  <div class="stats">
    <div class="stat"><div class="num" id="s-total">—</div><div class="lbl">Total Students</div></div>
    <div class="stat"><div class="num" id="s-active">—</div><div class="lbl">In Progress</div></div>
    <div class="stat"><div class="num" id="s-submitted">—</div><div class="lbl">Submitted</div></div>
    <div class="stat"><div class="num" id="s-terminated">—</div><div class="lbl">Terminated</div></div>
    <div class="stat"><div class="num" id="s-violations" style="color:#dc2626">—</div><div class="lbl">Total Violations</div></div>
  </div>
  <div class="card">
    <div class="card-hdr"><h2>👥 Students</h2><span class="note" id="refresh-ts">Auto-refresh every 10s</span></div>
    <table><thead><tr><th>Name</th><th>Student ID</th><th>Status</th><th>Score</th><th>Violations</th><th>Submitted At</th></tr></thead>
    <tbody id="t-students"></tbody></table>
  </div>
  <div class="card">
    <div class="card-hdr"><h2>🚨 Violation Log</h2><span class="note" id="event-count">0 events</span></div>
    <table><thead><tr><th>Time</th><th>Student</th><th>ID</th><th>Event</th><th>Details</th></tr></thead>
    <tbody id="t-events"></tbody></table>
  </div>
</div>
<script>
const EXAM_ID="{{ exam.id }}",TOKEN="{{ admin_token }}";
const EVENT_BADGE={TAB_SWITCH:'bg-yellow',WINDOW_BLUR:'bg-gray',FULLSCREEN_EXIT:'bg-yellow',COPY_ATTEMPT:'bg-blue',PASTE_ATTEMPT:'bg-blue',CUT_ATTEMPT:'bg-blue',KEY_BLOCKED:'bg-gray',PRINTSCREEN:'bg-blue',WINDOWS_KEY:'bg-blue',DEVTOOLS_OPEN:'bg-red',DEVTOOLS_ATTEMPT:'bg-red',COPILOT_ATTEMPT:'bg-red',ALT_TAB:'bg-yellow',RIGHT_CLICK:'bg-gray'};
async function loadDashboard(){
  const r=await fetch('/exam-api/exams/'+EXAM_ID+'/admin?token='+TOKEN);
  if(!r.ok){console.error('Failed');return;}
  const data=await r.json();
  const students=data.students,events=data.events;
  document.getElementById('s-total').textContent=students.length;
  document.getElementById('s-active').textContent=students.filter(s=>s.status==='in_progress').length;
  document.getElementById('s-submitted').textContent=students.filter(s=>s.status==='submitted').length;
  document.getElementById('s-terminated').textContent=students.filter(s=>s.status==='terminated').length;
  document.getElementById('s-violations').textContent=events.length;
  document.getElementById('event-count').textContent=events.length+' events';
  const tb=document.getElementById('t-students');
  tb.innerHTML=students.length===0?'<tr><td colspan="6" style="text-align:center;color:#94a3b8;padding:18px">No students yet.</td></tr>':students.map(s=>{const bc=s.status==='submitted'?'bg-green':s.status==='terminated'?'bg-red':'bg-blue';const scoreLabel=s.maxScore>0?`${s.score}/${s.maxScore}`:'—';return`<tr><td>${s.studentName}</td><td class="mono">${s.studentId}</td><td><span class="badge ${bc}">${s.status}</span></td><td>${scoreLabel}</td><td>${s.violationCount>0?'<span class="badge bg-red">'+s.violationCount+'</span>':'0'}</td><td>${s.submittedAt?new Date(s.submittedAt+'Z').toLocaleString():'—'}</td></tr>`;
      }).join('');

  const te = document.getElementById('t-events');
  te.innerHTML = events.length === 0
    ? '<tr><td colspan="5" style="text-align:center;color:#94a3b8;padding:18px">No violations logged yet.</td></tr>'
    : events.slice().reverse().map(e => {
        const bc = EVENT_BADGE[e.eventType] || 'bg-gray';
        const t = new Date(e.timestamp + 'Z').toLocaleTimeString();
        return `<tr>
          <td style="white-space:nowrap">${t}</td>
          <td>${e.studentName}</td>
          <td class="mono">${e.studentId}</td>
          <td><span class="badge ${bc} mono">${e.eventType}</span></td>
          <td style="color:#64748b">${e.details || ''}</td>
        </tr>`;
      }).join('');

  document.getElementById('refresh-ts').textContent = 'Last updated: ' + new Date().toLocaleTimeString();
}

function downloadXlsx() {
  window.location.href = '/exam-api/exams/' + EXAM_ID + '/export-xlsx?token=' + TOKEN;
}

loadDashboard();
setInterval(loadDashboard, 10000);
</script>
</body>
</html>
"""

@app.route("/")
def home():
    return render_template_string(HOME_TEMPLATE)

@app.route("/exam/<exam_id>")
def exam_register(exam_id):
    conn = get_db()
    exam = conn.execute("SELECT * FROM exams WHERE id = ?", (exam_id,)).fetchone()
    conn.close()
    if not exam: abort(404)
    return render_template_string(REGISTER_TEMPLATE, exam={
        "id": exam["id"], "title": exam["title"], "duration": exam["duration"]
    })

@app.route("/exam/<exam_id>/take")
def exam_take(exam_id):
    conn = get_db()
    exam = conn.execute("SELECT * FROM exams WHERE id = ?", (exam_id,)).fetchone()
    conn.close()
    if not exam: abort(404)
    questions = json.loads(exam["questions"])
    
    # Get session ID from query parameter to retrieve and shuffle for this student
    session_id = request.args.get('sessionId', '')
    if session_id:
        conn = get_db()
        session = conn.execute("SELECT * FROM sessions WHERE id = ?", (session_id,)).fetchone()
        conn.close()
        if session and not session["question_order"]:
            # First time this student is taking the exam - shuffle and store
            shuffled_indices = list(range(len(questions)))
            random.shuffle(shuffled_indices)
            question_order = json.dumps(shuffled_indices)
            conn = get_db()
            conn.execute("UPDATE sessions SET question_order = ? WHERE id = ?", (question_order, session_id))
            conn.commit()
            conn.close()
            questions = [questions[i] for i in shuffled_indices]
        elif session and session["question_order"]:
            # Already shuffled - use stored order
            shuffled_indices = json.loads(session["question_order"])
            questions = [questions[i] for i in shuffled_indices]
    
    return render_template_string(EXAM_TEMPLATE, exam={
        "id": exam["id"], "title": exam["title"], "duration": exam["duration"]
    }, questions=questions)

@app.route("/admin/<exam_id>")
def admin_view(exam_id):
    token = request.args.get("token", "")
    conn = get_db()
    exam = conn.execute("SELECT * FROM exams WHERE id = ?", (exam_id,)).fetchone()
    conn.close()
    if not exam: abort(404)
    if exam["admin_token"] != token: abort(403)
    return render_template_string(ADMIN_TEMPLATE, exam={
        "id": exam["id"], "title": exam["title"],
        "duration": exam["duration"], "teacher_email": exam["teacher_email"]
    }, admin_token=token)


@app.route("/exam-api/parse-docx", methods=["POST"])
def api_parse_docx():
    """Parse MCQs from an uploaded Word document."""
    if not DOCX_AVAILABLE:
        return jsonify({"error": "Word file support not available. Install: pip install python-docx"}), 400
    
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400
    
    file = request.files['file']
    if not file.filename.endswith('.docx'):
        return jsonify({"error": "File must be a .docx file"}), 400
    
    try:
        questions = parse_docx_questions(file)
        if not questions:
            return jsonify({"error": "No questions found in the document"}), 400
        return jsonify({"questions": questions}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/exam-api/exams", methods=["POST"])
def api_create_exam():
    data = request.get_json()
    if not data: return jsonify({"error": "Invalid JSON"}), 400
    for f in ["teacherEmail", "title", "duration", "questions"]:
        if f not in data: return jsonify({"error": f"Missing: {f}"}), 400
    exam_id = str(uuid.uuid4())[:8]
    admin_token = str(uuid.uuid4())
    now = datetime.utcnow().isoformat()
    conn = get_db()
    conn.execute(
        "INSERT INTO exams (id, teacher_email, title, duration, questions, admin_token, created_at) VALUES (?,?,?,?,?,?,?)",
        (exam_id, data["teacherEmail"], data["title"], int(data["duration"]),
         json.dumps(data["questions"]), admin_token, now)
    )
    conn.commit(); conn.close()
    base = request.host_url.rstrip("/")
    return jsonify({
        "examId": exam_id,
        "adminToken": admin_token,
        "examLink": f"{base}/exam/{exam_id}",
        "adminLink": f"{base}/admin/{exam_id}?token={admin_token}"
    }), 201


@app.route("/exam-api/exams/<exam_id>/students", methods=["POST"])
def api_register_student(exam_id):
    conn = get_db()
    exam = conn.execute("SELECT id FROM exams WHERE id = ?", (exam_id,)).fetchone()
    if not exam: conn.close(); return jsonify({"error": "Exam not found"}), 404
    data = request.get_json()
    if not data or not data.get("name") or not data.get("studentId"):
        conn.close(); return jsonify({"error": "name and studentId required"}), 400
    # Check if student is blocked
    blocked = conn.execute("SELECT id FROM sessions WHERE exam_id=? AND student_id=? AND status='terminated'", (exam_id, data["studentId"])).fetchone()
    if blocked:
        conn.close(); return jsonify({"error": "You have been blocked from this exam due to previous violations."}), 403
    session_id = str(uuid.uuid4())
    now = datetime.utcnow().isoformat()
    conn.execute(
        "INSERT INTO sessions (id, exam_id, student_name, student_id, status, created_at) VALUES (?,?,?,?,?,?)",
        (session_id, exam_id, data["name"], data["studentId"], "in_progress", now)
    )
    conn.commit(); conn.close()
    return jsonify({"sessionId": session_id, "studentName": data["name"], "studentId": data["studentId"]}), 201


@app.route("/exam-api/session/<session_id>")
def api_get_session(session_id):
    conn = get_db()
    s = conn.execute("SELECT * FROM sessions WHERE id = ?", (session_id,)).fetchone()
    conn.close()
    if not s: return jsonify({"error": "Not found"}), 404
    return jsonify({"sessionId": s["id"], "studentName": s["student_name"],
                    "studentId": s["student_id"], "status": s["status"]})


@app.route("/exam-api/exams/<exam_id>/submit", methods=["POST"])
def api_submit_exam(exam_id):
    data = request.get_json()
    if not data or not data.get("sessionId"):
        return jsonify({"error": "sessionId required"}), 400
    status = "terminated" if data.get("terminated") else "submitted"
    now = datetime.utcnow().isoformat()

    conn = get_db()
    # Load exam questions for scoring
    exam = conn.execute("SELECT * FROM exams WHERE id = ?", (exam_id,)).fetchone()
    questions = []
    if exam:
        try:
            questions = json.loads(exam["questions"])
        except Exception:
            questions = []

    answers = data.get("answers", {}) or {}
    if not isinstance(answers, dict):
        answers = {}

    score = 0
    max_score = 0
    for q in questions:
        marks = int(q.get("marks") or 0)
        max_score += marks
        if marks <= 0:
            continue
        qid = q.get("id")
        if not qid:
            continue
        if q.get("type") == "mcq":
            correct = str(q.get("correctAnswer", ""))
            if str(answers.get(qid, "")) == correct:
                score += marks
        else:
            correct = str(q.get("correctAnswer", "")).strip().lower()
            ans = str(answers.get(qid, "")).strip().lower()
            if correct and ans and ans == correct:
                score += marks

    conn.execute(
        "UPDATE sessions SET status=?, answers=?, submitted_at=?, score=?, max_score=? WHERE id=?",
        (status, json.dumps(answers), now, score, max_score, data["sessionId"])
    )
    conn.commit(); conn.close()
    return jsonify({"success": True, "message": "Submitted", "score": score, "maxScore": max_score})


@app.route("/exam-api/events", methods=["POST"])
def api_log_event():
    data = request.get_json()
    if not data: return jsonify({"success": False, "terminate": False}), 400
    exam_id    = data.get("examId", "")
    session_id = data.get("sessionId", "")
    event_type = data.get("eventType", "UNKNOWN")
    details    = data.get("details", "")
    now = datetime.utcnow().isoformat()
    conn = get_db()
    session = conn.execute("SELECT * FROM sessions WHERE id = ?", (session_id,)).fetchone()
    if session and session["status"] == "terminated":
        conn.close(); return jsonify({"success": True, "terminate": False})
    student_name = session["student_name"] if session else "Unknown"
    student_id   = session["student_id"]   if session else "—"
    conn.execute(
        "INSERT INTO events (exam_id, session_id, student_name, student_id, event_type, details, timestamp) VALUES (?,?,?,?,?,?,?)",
        (exam_id, session_id, student_name, student_id, event_type, details, now)
    )
    conn.commit()
    conn.close()
   
    return jsonify({"success": True, "terminate": False})


@app.route("/exam-api/exams/<exam_id>/admin")
def api_admin_data(exam_id):
    token = request.args.get("token", "")
    conn = get_db()
    exam = conn.execute("SELECT * FROM exams WHERE id = ?", (exam_id,)).fetchone()
    if not exam or exam["admin_token"] != token:
        conn.close(); return jsonify({"error": "Forbidden"}), 403
    sessions = conn.execute("SELECT * FROM sessions WHERE exam_id=? ORDER BY created_at", (exam_id,)).fetchall()
    events   = conn.execute("SELECT * FROM events WHERE exam_id=? ORDER BY timestamp", (exam_id,)).fetchall()
    conn.close()
    vc = {}
    for e in events: vc[e["session_id"]] = vc.get(e["session_id"], 0) + 1

    # Use stored score/max_score from sessions (added via schema migration)
    student_list = []
    for s in sessions:
        student_list.append({
            "sessionId": s["id"],
            "studentName": s["student_name"],
            "studentId": s["student_id"],
            "status": s["status"],
            "violationCount": vc.get(s["id"], 0),
            "submittedAt": s["submitted_at"],
            "score": s["score"] or 0,
            "maxScore": s["max_score"] or 0,
        })

    return jsonify({
        "exam": {"id": exam["id"], "title": exam["title"],
                 "duration": exam["duration"], "teacher_email": exam["teacher_email"]},
        "students": student_list,
        "events": [{"id": e["id"], "examId": e["exam_id"], "sessionId": e["session_id"],
                    "studentName": e["student_name"], "studentId": e["student_id"],
                    "eventType": e["event_type"], "details": e["details"] or "",
                    "timestamp": e["timestamp"]} for e in events]
    })
@app.route("/exam-api/exams/<exam_id>/export-xlsx")
def api_export_xlsx(exam_id):
    token = request.args.get("token", "")
    conn = get_db()
    exam = conn.execute("SELECT * FROM exams WHERE id = ?", (exam_id,)).fetchone()
    if not exam or exam["admin_token"] != token:
        conn.close(); return jsonify({"error": "Forbidden"}), 403
    events = conn.execute("SELECT * FROM events WHERE exam_id=? ORDER BY timestamp", (exam_id,)).fetchall()
    conn.close()
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Violations"
        headers = ["#", "Exam ID", "Exam Title", "Student Name", "Student ID",
                   "Event Type", "Details", "Timestamp", "Session ID"]
        hfill = PatternFill("solid", fgColor="1A56DB")
        hfont = Font(bold=True, color="FFFFFF")
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = hfill; cell.font = hfont
            cell.alignment = Alignment(horizontal="center")
        for ri, e in enumerate(events, 2):
            ws.cell(ri, 1, ri - 1)
            ws.cell(ri, 2, e["exam_id"])
            ws.cell(ri, 3, exam["title"])
            ws.cell(ri, 4, e["student_name"])
            ws.cell(ri, 5, e["student_id"])
            ws.cell(ri, 6, e["event_type"])
            ws.cell(ri, 7, e["details"] or "")
            ws.cell(ri, 8, e["timestamp"])
            ws.cell(ri, 9, e["session_id"])
        for i, w in enumerate([5, 12, 28, 24, 14, 22, 38, 24, 36], 1):
            ws.column_dimensions[chr(64 + i)].width = w
        out = BytesIO()
        wb.save(out); out.seek(0)
        return send_file(out,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"violations_{exam_id}.xlsx")
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500
if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", 5000))
    use_ssl = str(os.environ.get("USE_SSL", "")).lower() in ("1", "true", "yes")
    scheme = "https" if use_ssl else "http"
    print(f"\n{'='*60}")
    print("  ExamGuard – Online Exam Anti-Cheating System")
    print(f"{'='*60}")
    print(f"  URL:    {scheme}://localhost:{port}")
    print(f"  Routes: /exam-api/... (not /api/)")
    print(f"{'='*60}\n")
    app.run(host="0.0.0.0", port=port, debug=False, ssl_context="adhoc" if use_ssl else None)